from .base import FileLikeObject
from .custom_data import RecordedDataItem, RecordedDataType
from .experiment import ExperimentLevel
from .metadata import PictureMetadataPicturePlanes
from .nd2 import Nd2Reader

import base64, hashlib, itertools, json, os

def allImageInformationAsJsons(file_like: FileLikeObject, *, chunker_kwargs) -> tuple[str, str]:
    with Nd2Reader(file_like, chunker_kwargs=chunker_kwargs) as reader:
        return imageInformationAsJson(reader), recordedDataAsJson(reader)

def imageInformationAsJson(file_object : Nd2Reader) -> str:
    ret = {}
    ret = file_object.generalImageInfo
    ret["imageTextInfo"] = file_object.imageTextInfo.to_dict() if file_object.imageTextInfo is not None else {}

    exp_data = []
    if file_object.experiment is not None:
        for exp in file_object.experiment:
            json_ = json.dumps(_experiment_to_table(exp))
            exp_data.append(dict(ClassName=exp.name.lower(), LoopName=f'{exp.name} Loop', InfoData="data:application/json;base64," + base64.b64encode(json_.encode()).decode()))
    ret["experimentData"] = exp_data

    custom_desc = []
    if file_object.customDescription is not None:
        for item in file_object.customDescription:
            custom_desc.append(dict(name=item.name, text=item.valueAsText, type=int(item.type)))
    ret["customMetadata"] = custom_desc

    json_ = json.dumps(_picture_planes_to_table(file_object.pictureMetadata.sPicturePlanes))
    ret["acquisitionDetails"] = "data:application/json;base64," + base64.b64encode(json_.encode()).decode()
    return json.dumps(ret)

def recordedDataAsJson(file_object : Nd2Reader) -> str:
    return json.dumps(_recorded_data_to_table(file_object.recordedData))

def imageDataAsDict(file_object: Nd2Reader) -> str:
    ret = {}
    ret["generalInfo"] = file_object.generalImageInfo
    ret["imageTextInfo"] = file_object.imageTextInfo.to_dict() if file_object.imageTextInfo is not None else {}

    exp_data = []
    if file_object.experiment is not None:
        for exp in file_object.experiment:
            json_ = _experiment_to_table(exp)
            exp_data.append(dict(ClassName=exp.name.lower(), LoopName=f'{exp.name} Loop', data=json_))
    ret["experimentData"] = exp_data

    custom_desc = []
    if file_object.customDescription is not None:
        for item in file_object.customDescription:
            custom_desc.append(dict(name=item.name, text=item.valueAsText, type=int(item.type)))
    ret["customMetadata"] = custom_desc
    ret["recordedData"] = _recorded_data_to_table(file_object.recordedData)
    ret["acquisitionDetails"] = _picture_planes_to_table(file_object.pictureMetadata.sPicturePlanes)
    return ret

def _get_format_fn(val) -> str:
    return """(coldef) => {
        coldef._number_format = Intl.NumberFormat(navigator.language, { minimumFractionDigits : %u, maximumFractionDigits: %u });
        coldef.fmtfn = function(val) {
            return this._number_format.format(val);
        };
        coldef.fmtfn.bind(coldef);
    };""" % (val, val)

def _create_treeview_grouping(rows: list[dict[str, any]], groupby: list[str], ordering: dict[dict[str, int]]|None = None) -> list[dict[str, any]]:
    sort_keys = {}
    for grpcol in groupby:
        if ordering is not None and grpcol in ordering:
            sort_keys[grpcol] = lambda row: ordering[grpcol][row[grpcol]]
        else:
            sort_keys[grpcol] = lambda row: row[grpcol]

    def recursive_fn(parent: dict[str, any], hash: object, rows: list[dict[str, any]], groupby: list[str], depth: int):
        rcount = 0
        gcount = 0
        ret_group_list = []
        grpcol = groupby.pop(0)
        for k, g in itertools.groupby(rows, key=lambda row: row[grpcol]):
            grouprows = list(g)
            start = rows.index(grouprows[0])
            end = start + len(grouprows)
            hh = hash.copy()
            hh.update(json.dumps(k).encode())
            group = dict(id=hh.hexdigest(), title=str(k), colid=grpcol, depth=depth, groupcount=0, rowcount=end-start, rows=(start, end))
            ret_group_list.append(group)
            rcount += end - start
            gcount += 1
            if len(groupby):
                ret_group_list += recursive_fn(group, hh, rows, groupby, depth+1)
        parent["rowcount"] -= rcount
        parent["groupcount"] = gcount
        return ret_group_list

    for grpcol in reversed(groupby):
        rows.sort(key=sort_keys[grpcol])

    h = hashlib.sha256()
    h.update(b"root")
    rowcount = len(rows)
    root = dict(id=h.hexdigest(), title="All", depth=0, groupcount=0, rowcount=rowcount, rows=(0, rowcount))
    ret = [ root ]
    ret += recursive_fn(root, h, rows, groupby, 1)
    return ret

def _experiment_to_table(exp: ExperimentLevel) -> dict[str, any]:
    right_align = {'text-align': 'right'}
    css_style = { 'X': right_align, 'Y': right_align, 'Z': right_align, 'Bottom': right_align, 'Count': right_align, 'Step': right_align, 'Top': right_align, 'Interval': right_align, 'Duration': right_align, 'Loops': right_align }
    min_width = { 'X': '100px', 'Y': '100px', 'Z': '100px', 'Bottom': '80px', 'Count': '60px', 'Step': '80px', 'Top': '80px', 'Interval': '100px', 'Duration': '100px', 'Loops': '100px' }
    format_fn = { 'X': _get_format_fn(2), 'Y': _get_format_fn(2), 'Z': _get_format_fn(3), 'Bottom': _get_format_fn(3), 'Step': _get_format_fn(3), 'Top': _get_format_fn(3), 'Color': '(coldef) => { coldef.fmtfn = function(val) { return val === "#ffffff" ? "Brightfield" : "" }; };' }
    style_fn = { 'Color':  '(coldef) => { coldef.stylefn = function(val) { return  val === "#ffffff" ? { "background": "linear-gradient(0.25turn, rgba(255,0,0,0.3), rgba(0,255,0,0.3), rgba(0,0,255,0.3))" } : { "background-color": `${val}ee` }; } };' }
    replace = { 'X': 'X Pos [µm]', 'Y': 'Y Pos [µm]', 'Z': 'Z Pos [µm]', 'OC': 'Opt. conf.', 'Bottom': 'Bottom [µm]', 'Drive': 'Z Drive', 'Step': 'Z Step [µm]', 'Top': 'Top [µm]' }
    col_defs = []
    for k in exp.uLoopPars.info[0].keys():
        s = css_style.get(k, {})
        d = dict(id=k, title=replace.get(k, k), headerStyle=s, style=s)
        fmt_fn_code = format_fn.get(k, None)
        if fmt_fn_code is not None:
            d["fmtfncode"] = fmt_fn_code
        style_fn_code = style_fn.get(k, None)
        if style_fn_code is not None:
            d["stylefncode"] = style_fn_code
        min_w = min_width.get(k, None)
        if min_w is not None:
            d["minwidth"] = min_w
        col_defs.append(d)

    return dict(coldefs=col_defs, rowdata=exp.uLoopPars.info)

def _picture_planes_to_table(planes: PictureMetadataPicturePlanes) -> dict[str, any]:
    rows=[]
    col_defs=[ dict(id="id", hidden=True), dict(id="camera", title="Camera"), dict(id="channel", title="Channel"), dict(id="feature", title="Feature"), dict(id="value", title="Value") ]
    settings = planes.sSampleSetting
    for plane in planes.sPlaneNew:
        setting = settings[plane.uiSampleIndex] if 0 <= plane.uiSampleIndex < len(settings) else None
        if setting:
            camera = setting.cameraName or "Unknown camera"
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="OC name:", value=','.join(oc for oc in setting.opticalConfigurations)))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Microscope name:", value=setting.microscopeName))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective name:", value=setting.objectiveName))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective magnification:", value=setting.objectiveMagnification))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective numerical aperture:", value=setting.objectiveNumericAperture))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Refractive index:", value=setting.refractiveIndex))
        else:
            camera = "Unknown camera"
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="OC name:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Microscope name:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective name:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective magnification:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective numerical aperture:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Refractive index:", value='N/A'))
    rows.sort(key=lambda row: row["camera"])
    groupedBy = ['camera', 'channel']
    d = dict(coldefs=col_defs, groups=_create_treeview_grouping(rows, groupedBy.copy()), rowdata=rows, groupedby=groupedBy)
    return d

def _get_recorded_data_fmt_function(col: RecordedDataItem) -> str:
    if col.Type == RecordedDataType.eDouble:
        digits = 2 if col.ID in ('X', 'Y') else 3
        return _get_format_fn(digits)
    else:
        return "(coldef) => { coldef.fmtfn = String };"

def _get_recorded_data_styles(col: RecordedDataItem) -> dict[str, str]:
    if col.Type in (RecordedDataType.eDouble, RecordedDataType.eInt) or col.ID == "ACQTIME":
        return { "text-align": "right" }
    else:
        return { "text-align": "left" }

def _recorded_data_to_table(recdata: RecordedDataType) -> dict[str, any]:
    coldefs = []
    coldefs.append(dict(id='id', hidden=True))
    rowdata = [ dict(id=i+1) for i in range(recdata.rowCount) ]
    for col in recdata.data:
        coldefs.append(dict(id=col.ID, title=f"{col.Desc} [{col.Unit}]" if col.Unit else col.Desc, fmtfncode=_get_recorded_data_fmt_function(col), style=_get_recorded_data_styles(col)))
        for index, datavalue in enumerate(col.data):
            rowdata[index][col.ID] = datavalue
    coldefs.append(dict(id='tail'))
    return dict(coldefs=coldefs, rowdata=rowdata)

def table_to_TSV(table, header=True, include_all_columns=False):
    rootGroupId = "4813494d137e1631bba301d5acab6e7bb7aa74ce1185d456565ef51d737677b2"

    coldefs = table["coldefs"]
    items = table["rowdata"]
    groups = table.get("groups", [{
        "id": rootGroupId, "depth": 0, "groupcount": 0, "rowcount": len(items), "rows": [0, len(items)]
    }])

    groupedColumns = table.get("groupedby", [])

    rootGroupId = "4813494d137e1631bba301d5acab6e7bb7aa74ce1185d456565ef51d737677b2"

    export_data = ""

    if header:
        for coldef in coldefs:
            if (include_all_columns or (not coldef.get("hidden", False))) and (coldef["id"] not in groupedColumns):
                export_data += (coldef.get("title", "") or "") + "\t"
        export_data = export_data.strip() + "\n"

    def add_group_title(group, depth):
        if group["id"] == rootGroupId:
            return ""
        return "\t" * (depth - 1) + f"{group['title']}\n"

    def add_item_data(item, depth):
        row = "\t" * depth
        for coldef in coldefs:
            if (include_all_columns or (not coldef.get("hidden", False))) and (coldef["id"] not in groupedColumns):
                data = item.get(coldef["id"], "")
                item_text = coldef.get("fmtfn", lambda x: x)(data) if data else ""
                row += str(item_text).strip() + "\t"
        return row + "\n"

    def add_group_data(group, depth):
        group_data = add_group_title(group, depth)
        group_items = items[group["rows"][0]:group["rows"][1]]

        for item in group_items:
            group_data += add_item_data(item, depth)

        child_groups = [g for g in groups if g.get("parentId") == group["id"]]
        for child_group in child_groups:
            group_data += add_group_data(child_group, depth + 1)

        return group_data

    pxcount = 0
    skip_deeper = float("inf")

    for group in groups:
        if skip_deeper < group["depth"]:
            continue
        skip_deeper = group.get("open", True) and float("inf") or group["depth"]

        if group.get("visible", True):
            export_data += add_group_title(group, group["depth"])

        if group["groupcount"] == 0 and group["depth"] < skip_deeper:
            for j in range(group["rowcount"]):
                export_data += add_item_data(items[group["rows"][0] + j], group["depth"])

    return export_data

def maybe_wrap_field(value):
    if value is None:
        return ""
    if '\n' in value:
        return f"{value.replace('\n', '\n\t')}"
    return value

def export_main_image_info(image_info):
    return "\n".join([
        f"Filename:\t{maybe_wrap_field(image_info['generalInfo'].get('filename', ""))}",
        f"Path:\t{maybe_wrap_field(image_info['generalInfo'].get('path', ""))}",
        f"Dimension:\t{maybe_wrap_field(image_info['generalInfo'].get('dimension', ""))}",
        f"Sizes:\t{maybe_wrap_field(image_info['generalInfo'].get('sizes', ""))}",
        f"Modified time:\t{maybe_wrap_field(image_info['generalInfo'].get('mtime', ""))}",
        f"Created by:\t{maybe_wrap_field(image_info['generalInfo'].get('app_created', ""))}"
    ])

def export_image_text_info(image_info):
    return "\n".join([
        f"Calibration:\t{maybe_wrap_field(image_info['generalInfo'].get('calibration', ""))}",
        f"Optics:\t{maybe_wrap_field(image_info['imageTextInfo'].get('optics', ''))}",
        f"Type:\t{maybe_wrap_field(image_info['imageTextInfo'].get('type', ''))}",
        f"Sample ID:\t{maybe_wrap_field(image_info['imageTextInfo'].get('sampleId', ''))}",
        f"Author:\t{maybe_wrap_field(image_info['imageTextInfo'].get('author', ''))}",
        f"Description:\t{maybe_wrap_field(image_info['imageTextInfo']['description'])}",
        f"Image ID:\t{maybe_wrap_field(image_info['imageTextInfo'].get('imageId', ''))}",
        f"Group:\t{maybe_wrap_field(image_info['imageTextInfo'].get('group', ''))}",
        f"Capturing:\t{maybe_wrap_field(image_info['imageTextInfo']['capturing'])}",
        f"Sampling:\t{maybe_wrap_field(image_info['imageTextInfo'].get('sampling', ''))}",
        f"Location:\t{maybe_wrap_field(image_info['imageTextInfo'].get('location', ''))}",
        f"Date:\t{maybe_wrap_field(image_info['imageTextInfo'].get('date', ''))}",
        f"Conclusion:\t{maybe_wrap_field(image_info['imageTextInfo'].get('conclusion', ''))}",
        f"Info 1:\t{maybe_wrap_field(image_info['imageTextInfo'].get('info1', ''))}",
        f"Info 2:\t{maybe_wrap_field(image_info['imageTextInfo'].get('info2', ''))}"
    ])

def export_experiments(image_info):
    output = ""
    if image_info.get('experimentData') and len(image_info['experimentData']) > 0:
        for exp_data in image_info['experimentData']:
            output += f"{maybe_wrap_field(exp_data['LoopName'])}\n"
            output += table_to_TSV(exp_data['data'], header=True)
    return output

def export_recorded_data(image_info):
    output = ""
    if image_info.get('recordedData'):
        output += table_to_TSV(image_info['recordedData'], header=True)
    return output

def export_custom_metadata(image_info):
    output = ""
    if image_info.get('customMetadata') and len(image_info['customMetadata']) > 0:
        for item in image_info['customMetadata']:
            output += f"{maybe_wrap_field(item['name'])}\t{maybe_wrap_field(item['text'])}\n"
    return output

def export_acquisition_details(image_info):
    output = ""
    if image_info.get('acquisitionDetails'):
        output += table_to_TSV(image_info['acquisitionDetails'], header=False)
    return output


def imageInformationAsTSV(filename):
    data = imageDataAsDict(filename)

    result = ""
    result += "Main Image Info:\n" + export_main_image_info(data) + "\n\n"
    result += "Image Text Info:\n" + export_image_text_info(data) + "\n\n"
    if data.get('experimentData') and len(data['experimentData']) > 0:
        result += "Experiments:\n" + export_experiments(data) + "\n\n"
    if data.get('customMetadata') and len(data['customMetadata']) > 0:
        result += "Custom Metadata:\n" + export_custom_metadata(data) + "\n\n"
    if data.get('recordedData'):
        result += "Recorded Data:\n" + export_recorded_data(data) + "\n\n"
    if data.get('acquisitionDetails'):
        result += "Acquisition Details:\n" + export_acquisition_details(data) + "\n\n"
    return result

def imageInformationAsXLSL(filename):

    import io as IO

    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    data = imageDataAsDict(filename)
    def add_sheet(sheet_name, data, col_widths=None, row_height=15):
        ws = wb.create_sheet(title=sheet_name)
        rows = [row.split('\t') for row in data.strip().split('\n')]

        for row_idx, row in enumerate(rows, start=1):
            ws.row_dimensions[row_idx].height = row_height
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell.replace('\n', '').replace('\r', ''))

        if isinstance(col_widths, list):
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        elif isinstance(col_widths, int):
            max_cols = max(len(row) for row in rows)
            for col_idx in range(1, max_cols + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths

    sections = {
        "Main Image Info": (export_main_image_info, [25, 100], 15),
        "Image Text Info": (export_image_text_info, [25, 100], 15),
        "Experiment Data": (export_experiments, 15, 15),
        "Recorded Data": (export_recorded_data, 15, 15),
        "Custom Metadata": (export_custom_metadata, [25, 100], 15),
        "Acquisition Details": (export_acquisition_details, 20, 15)
    }

    for sheet_name, (export_func, col_width, row_height) in sections.items():
        add_sheet(sheet_name, export_func(data), col_width, row_height)

    wb.remove(wb["Sheet"])
    result = IO.BytesIO()
    wb.save(result)
    result.seek(0)
    return base64.b64encode(result.getvalue()).decode('utf-8')



def format_file_size(size_bytes: int) -> str:
    """Converts file size to human-readable format."""
    units = ["B", "KB", "MB", "GB", "TB"]
    size = float(size_bytes)
    unit_index = 0
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    return f"{size:.2f} {units[unit_index]}"

def get_nonND2_image_info(file_path: str) -> dict[str, any]:
    import matplotlib.pyplot as plt
    from datetime import datetime
    """
    Retrieves general information about an image file using only os and matplotlib. (for non-nd2 files)
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    filename = os.path.basename(file_path)
    path = os.path.dirname(file_path)
    file_size = os.path.getsize(file_path)
    mtime = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%m/%d/%y %H:%M:%S')

    try:
        img = plt.imread(file_path)
        height, width = img.shape[:2]
        bit_depth = img.dtype.itemsize * 8
        num_channels = img.shape[2] if len(img.shape) == 3 else 1
    except Exception as e:
        height, width, bit_depth, num_channels = None, None, None, None

    return {
        "filename": filename,
        "path": path,
        "sizes": format_file_size(file_size),
        "dimension": f"{width} x {height}" if width and height else "Unknown",
        "bit_depth": f"{bit_depth}-bit" if bit_depth else "Unknown",
        "num_channels": num_channels if num_channels else "Unknown",
        "mtime": mtime,
        "nonND2": True
    }