import os
from typing import Any
from .attributes import ImageAttributesPixelType
from .base import FileLikeObject
from .custom_data import RecordedData, RecordedDataItem, RecordedDataType
from .experiment import ExperimentLevel
from .metadata import PictureMetadataPicturePlanes
from .nd2 import Nd2Reader

import hashlib, itertools, json

def generalImageInfo(reader: Nd2Reader) -> dict[str, Any]:
    """
    Returns general information about the image as a dictionary.
    """
    ia = reader.imageAttributes
    loops = ", ".join([ f"{exp_level.shortName}({exp_level.count})" for exp_level in reader.experiment if 0 < exp_level.count ]) if reader.experiment else ""
    path = ""
    filename = ""
    if reader.store.filename:
        path, filename = os.path.split(reader.store.filename)
    elif reader.store.uri:
        path, filename = os.path.split(reader.store.uri.rstrip("/"))
    path += os.sep

    bit_depth = f"{ia.uiBpcSignificant}bit {ImageAttributesPixelType.short_name(ia.ePixelType)}"
    frame_res = f"{ia.width} x {ia.height}"
    dimension = f"{frame_res} ({ia.componentCount} {"comps" if 1 < ia.componentCount else "comp"} {bit_depth})" + (f" x {ia.uiSequenceCount} frames" if 1 < ia.uiSequenceCount else "") +(f": {loops}" if loops else "")
    calibration = f"{reader.pictureMetadata.dCalibration:.3f} µm/px" if reader.pictureMetadata.bCalibrated else "Uncalibrated"

    mtime = f"{reader.store.lastModified.strftime('%x %X')}"
    app_created = reader.appInfo.software

    sizes = format_general_info_sizes(reader.store.sizeOnDisk, ia.widthBytes*ia.height, ia.widthBytes*ia.height*reader.experiment.dims.get('z', 0) if reader.experiment is not None else 0)

    return dict(filename=filename, path=path, bit_depth=bit_depth, loops=loops, dimension=dimension, calibration=calibration, mtime=mtime, app_created=app_created, **sizes)

def imageInformationAsJSON(file_like: FileLikeObject, *, filename: str|None = None, last_modified: str|None = None) -> str:
    return json.dumps(gatherImageInformation(file_like, filename=filename, last_modified=last_modified))

def gatherImageInformation(file_like: FileLikeObject, *, filename: str|None = None, last_modified: str|None = None) -> dict[str, Any]:
    with Nd2Reader(file_like, chunker_kwargs=dict(filename=filename, last_modified=last_modified)) as reader:
        return gatherImageInfoFromNd2(reader)

def gatherImageInfoFromNd2(file_object: Nd2Reader) -> dict[str, Any]:
    ret = {}
    ret["generalInfo"] = generalImageInfo(file_object)
    ret["imageTextInfo"] = file_object.imageTextInfo.to_dict() if file_object.imageTextInfo is not None else {}

    exp_data = []
    if file_object.experiment is not None:
        for exp in file_object.experiment:
            json_ = _experiment_to_table(exp)
            exp_data.append(dict(ClassName=exp.name.lower(), LoopName=f'{exp.name} Loop', data=json_))
    ret["experimentData"] = exp_data

    custom_desc = []
    if file_object.customDescription is not None:
        for item in file_object.customDescription:
            custom_desc.append(dict(name=item.name, text=item.valueAsText, type=int(item.type)))
    ret["customMetadata"] = custom_desc
    ret["recordedData"] = _recorded_data_to_table(file_object.recordedData)
    ret["acquisitionDetails"] = _picture_planes_to_table(file_object.pictureMetadata.sPicturePlanes)
    return ret

def _get_format_fn(val) -> str:
    return """(coldef) => {
        coldef._number_format = Intl.NumberFormat(navigator.language, { minimumFractionDigits : %u, maximumFractionDigits: %u });
        coldef.fmtfn = function(val) {
            return this._number_format.format(val);
        };
        coldef.fmtfn.bind(coldef);
    };""" % (val, val)

def _create_treeview_grouping(rows: list[dict[str, Any]], groupby: list[str], ordering: dict[str, dict[str, int]]|None = None) -> list[dict[str, Any]]:
    sort_keys = {}
    for grpcol in groupby:
        if ordering is not None and grpcol in ordering:
            sort_keys[grpcol] = lambda row, grpcol=grpcol: ordering[grpcol][row[grpcol]]
        else:
            sort_keys[grpcol] = lambda row, grpcol=grpcol: row[grpcol]

    def recursive_fn(parent: dict[str, Any], hash: "hashlib._Hash", rows: list[dict[str, Any]], groupby: list[str], depth: int):
        rcount = 0
        gcount = 0
        ret_group_list = []
        grpcol = groupby[0]
        next_cols = groupby[1:]
        for k, g in itertools.groupby(rows, key=lambda row: row[grpcol]):
            grouprows = list(g)

            parent_start = parent["rows"][0]
            offset = rows.index(grouprows[0])
            start = parent_start + offset
            end = start + len(grouprows)

            hh = hash.copy()
            hh.update(json.dumps(k).encode())
            group = dict(id=hh.hexdigest(), title=str(k), colid=grpcol, depth=depth, groupcount=0, rowcount=end-start, rows=(start, end))
            ret_group_list.append(group)
            rcount += end - start
            gcount += 1
            if next_cols:
                ret_group_list += recursive_fn(group, hh, grouprows, next_cols, depth+1)
        parent["rowcount"] -= rcount
        parent["groupcount"] = gcount
        return ret_group_list

    for grpcol in reversed(groupby):
        rows.sort(key=sort_keys[grpcol])

    h = hashlib.sha256()
    h.update(b"root")
    rowcount = len(rows)
    root = dict(id=h.hexdigest(), title="All", depth=0, groupcount=0, rowcount=rowcount, rows=(0, rowcount))
    ret = [ root ]
    ret += recursive_fn(root, h, rows, groupby.copy(), 1)
    return ret

def _experiment_to_table(exp: ExperimentLevel) -> dict[str, Any]:
    right_align = {'text-align': 'right'}
    css_style = { 'X': right_align, 'Y': right_align, 'Z': right_align, 'Bottom': right_align, 'Count': right_align, 'Step': right_align, 'Top': right_align, 'Interval': right_align, 'Duration': right_align, 'Loops': right_align }
    min_width = { 'X': '100px', 'Y': '100px', 'Z': '100px', 'Bottom': '80px', 'Count': '60px', 'Step': '80px', 'Top': '80px', 'Interval': '100px', 'Duration': '100px', 'Loops': '100px' }
    format_fn = { 'X': _get_format_fn(2), 'Y': _get_format_fn(2), 'Z': _get_format_fn(3), 'Bottom': _get_format_fn(3), 'Step': _get_format_fn(3), 'Top': _get_format_fn(3), 'Color': '(coldef) => { coldef.fmtfn = function(val) { return val === "#ffffff" ? "Brightfield" : "" }; };' }
    style_fn = { 'Color':  '(coldef) => { coldef.stylefn = function(val) { return  val === "#ffffff" ? { "background": "linear-gradient(0.25turn, rgba(255,0,0,0.3), rgba(0,255,0,0.3), rgba(0,0,255,0.3))" } : { "background-color": `${val}ee` }; } };' }
    replace = { 'X': 'X Pos [µm]', 'Y': 'Y Pos [µm]', 'Z': 'Z Pos [µm]', 'OC': 'Opt. conf.', 'Bottom': 'Bottom [µm]', 'Drive': 'Z Drive', 'Step': 'Z Step [µm]', 'Top': 'Top [µm]' }
    col_defs = []
    for k in exp.uLoopPars.info[0].keys():
        s = css_style.get(k, {})
        d = dict(id=k, title=replace.get(k, k), headerStyle=s, style=s)
        fmt_fn_code = format_fn.get(k, None)
        if fmt_fn_code is not None:
            d["fmtfncode"] = fmt_fn_code
        style_fn_code = style_fn.get(k, None)
        if style_fn_code is not None:
            d["stylefncode"] = style_fn_code
        min_w = min_width.get(k, None)
        if min_w is not None:
            d["minwidth"] = min_w
        col_defs.append(d)

    return dict(coldefs=col_defs, rowdata=exp.uLoopPars.info)

def _picture_planes_to_table(planes: PictureMetadataPicturePlanes) -> dict[str, Any]:
    rows=[]
    col_defs=[ dict(id="id", hidden=True), dict(id="camera", title="Camera"), dict(id="channel", title="Channel"), dict(id="feature", title="Feature"), dict(id="value", title="Value") ]
    settings = planes.sSampleSetting
    for plane in planes.sPlaneNew:
        setting = settings[plane.uiSampleIndex] if 0 <= plane.uiSampleIndex < len(settings) else None
        if setting:
            camera = setting.cameraName or "Unknown camera"
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="OC name:", value=','.join(oc for oc in setting.opticalConfigurations)))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Microscope name:", value=setting.microscopeName))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective name:", value=setting.objectiveName))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective magnification:", value=setting.objectiveMagnification))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective numerical aperture:", value=setting.objectiveNumericAperture))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Refractive index:", value=setting.refractiveIndex))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Emission wavelength:", value=(plane.emissionWavelengthNm if plane.emissionWavelengthNm else 'N/A')))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Excitation wavelength:", value=(plane.excitationWavelengthNm if plane.excitationWavelengthNm else 'N/A')))
        else:
            camera = "Unknown camera"
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="OC name:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Microscope name:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective name:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective magnification:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Objective numerical aperture:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Refractive index:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Emission wavelength:", value='N/A'))
            rows.append(dict(id=str(len(rows)+1), camera=camera, channel=plane.sDescription, feature="Excitation wavelength:", value='N/A'))
    rows.sort(key=lambda row: row["camera"])
    groupedBy = ['camera', 'channel']
    d = dict(coldefs=col_defs, groups=_create_treeview_grouping(rows, groupedBy.copy()), rowdata=rows, groupedby=groupedBy)
    return d

def _get_recorded_data_fmt_function(col: RecordedDataItem) -> str:
    if col.Type == RecordedDataType.eDouble:
        digits = 2 if col.ID in ('X', 'Y') else 3
        return _get_format_fn(digits)
    else:
        return "(coldef) => { coldef.fmtfn = String };"

def _get_recorded_data_styles(col: RecordedDataItem) -> dict[str, str]:
    if col.Type in (RecordedDataType.eDouble, RecordedDataType.eInt) or col.ID == "ACQTIME":
        return { "text-align": "right" }
    else:
        return { "text-align": "left" }

def _recorded_data_to_table(recdata: RecordedData) -> dict[str, Any]:
    coldefs = []
    coldefs.append(dict(id='id', hidden=True))
    rowdata = [ dict(id=i+1) for i in range(recdata.rowCount) ]
    for col in recdata.data:
        coldefs.append(dict(id=col.ID, title=f"{col.Desc} [{col.Unit}]" if col.Unit else col.Desc, fmtfncode=_get_recorded_data_fmt_function(col), style=_get_recorded_data_styles(col)))
        for index, datavalue in enumerate(col.data):
            rowdata[index][col.ID] = datavalue
    coldefs.append(dict(id='tail'))
    return dict(coldefs=coldefs, rowdata=rowdata)

def table_to_TSV(table, header=True, include_all_columns=False):
    rootGroupId = "4813494d137e1631bba301d5acab6e7bb7aa74ce1185d456565ef51d737677b2"

    coldefs = table["coldefs"]
    items = table["rowdata"]
    groups = table.get("groups", [{
        "id": rootGroupId, "depth": 0, "groupcount": 0, "rowcount": len(items), "rows": [0, len(items)]
    }])

    groupedColumns = table.get("groupedby", [])

    rootGroupId = "4813494d137e1631bba301d5acab6e7bb7aa74ce1185d456565ef51d737677b2"

    export_data = ""

    if header:
        for coldef in coldefs:
            if (include_all_columns or (not coldef.get("hidden", False))) and (coldef["id"] not in groupedColumns):
                export_data += (coldef.get("title", "") or "") + "\t"
        export_data = export_data.strip() + "\n"

    def add_group_title(group, depth):
        if group["id"] == rootGroupId:
            return ""
        return "\t" * (depth - 1) + f"{group['title']}\n"

    def add_item_data(item, depth):
        row = "\t" * depth
        for coldef in coldefs:
            if (include_all_columns or (not coldef.get("hidden", False))) and (coldef["id"] not in groupedColumns):
                data = item.get(coldef["id"], "")
                item_text = coldef.get("fmtfn", lambda x: x)(data) if data else ""
                row += str(item_text).strip() + "\t"
        return row + "\n"

    def add_group_data(group, depth):
        group_data = add_group_title(group, depth)
        group_items = items[group["rows"][0]:group["rows"][1]]

        for item in group_items:
            group_data += add_item_data(item, depth)

        child_groups = [g for g in groups if g.get("parentId") == group["id"]]
        for child_group in child_groups:
            group_data += add_group_data(child_group, depth + 1)

        return group_data

    pxcount = 0
    skip_deeper = float("inf")

    for group in groups:
        if skip_deeper < group["depth"]:
            continue
        skip_deeper = group.get("open", True) and float("inf") or group["depth"]

        if group.get("visible", True):
            export_data += add_group_title(group, group["depth"])

        if group["groupcount"] == 0 and group["depth"] < skip_deeper:
            for j in range(group["rowcount"]):
                export_data += add_item_data(items[group["rows"][0] + j], group["depth"])

    return export_data

def maybe_wrap_field(value):
    if value is None:
        return ""
    if '\n' in value:
        return f"{value.replace('\n', '\n\t')}"
    return value

def export_main_image_info(image_info):
    gi = image_info.get('generalInfo', {})
    return "\n".join([
        f"Filename:\t{maybe_wrap_field(gi.get('filename', ""))}",
        f"Path:\t{maybe_wrap_field(gi.get('path', ""))}",
        f"Dimension:\t{maybe_wrap_field(gi.get('dimension', ""))}",
        f"Sizes:\t{maybe_wrap_field(gi.get('sizes', ""))}",
        f"Modified time:\t{maybe_wrap_field(gi.get('mtime', ""))}",
        f"Created by:\t{maybe_wrap_field(gi.get('app_created', ""))}"
    ])

def export_image_text_info(image_info):
    ret = []
    gi = image_info.get('generalInfo', {})
    ret.append(f"Calibration:\t{maybe_wrap_field(gi.get('calibration', "Uncalibrated"))}")

    ti = image_info.get('imageTextInfo', {})
    ret.append(f"Optics:\t{maybe_wrap_field(ti.get('optics', ''))}")
    ret.append(f"Type:\t{maybe_wrap_field(ti.get('type', ''))}")
    ret.append(f"Sample ID:\t{maybe_wrap_field(ti.get('sampleId', ''))}")
    ret.append(f"Author:\t{maybe_wrap_field(ti.get('author', ''))}")
    ret.append(f"Description:\t{maybe_wrap_field(ti.get('description', ''))}")
    ret.append(f"Image ID:\t{maybe_wrap_field(ti.get('imageId', ''))}")
    ret.append(f"Group:\t{maybe_wrap_field(ti.get('group', ''))}")
    ret.append(f"Capturing:\t{maybe_wrap_field(ti.get('capturing', ''))}")
    ret.append(f"Sampling:\t{maybe_wrap_field(ti.get('sampling', ''))}")
    ret.append(f"Location:\t{maybe_wrap_field(ti.get('location', ''))}")
    ret.append(f"Date:\t{maybe_wrap_field(ti.get('date', ''))}")
    ret.append(f"Conclusion:\t{maybe_wrap_field(ti.get('conclusion', ''))}")
    ret.append(f"Info 1:\t{maybe_wrap_field(ti.get('info1', ''))}")
    ret.append(f"Info 2:\t{maybe_wrap_field(ti.get('info2', ''))}")

    return "\n".join(ret)

def export_experiments(image_info):
    output = ""
    if image_info.get('experimentData') and len(image_info['experimentData']) > 0:
        for exp_data in image_info['experimentData']:
            output += f"{maybe_wrap_field(exp_data['LoopName'])}\n"
            output += table_to_TSV(exp_data['data'], header=True)
    return output

def export_recorded_data(image_info):
    output = ""
    if image_info.get('recordedData'):
        output += table_to_TSV(image_info['recordedData'], header=True)
    return output

def export_custom_metadata(image_info):
    output = ""
    if image_info.get('customMetadata') and len(image_info['customMetadata']) > 0:
        for item in image_info['customMetadata']:
            output += f"{maybe_wrap_field(item['name'])}\t{maybe_wrap_field(item['text'])}\n"
    return output

def export_acquisition_details(image_info):
    output = ""
    if image_info.get('acquisitionDetails'):
        output += table_to_TSV(image_info['acquisitionDetails'], header=False)
    return output


def imageInformationAsTXT(image_information: str|dict[str, Any]) -> str:
    data = json.loads(image_information) if isinstance(image_information, str) else image_information

    result = ""
    result += "Main Image Info:\n" + export_main_image_info(data) + "\n\n"
    result += "Image Text Info:\n" + export_image_text_info(data) + "\n\n"
    if data.get('experimentData') and len(data['experimentData']) > 0:
        result += "Experiments:\n" + export_experiments(data) + "\n\n"
    if data.get('customMetadata') and len(data['customMetadata']) > 0:
        result += "Custom Metadata:\n" + export_custom_metadata(data) + "\n\n"
    if data.get('recordedData'):
        result += "Recorded Data:\n" + export_recorded_data(data) + "\n\n"
    if data.get('acquisitionDetails'):
        result += "Acquisition Details:\n" + export_acquisition_details(data) + "\n\n"
    return result

def imageInformationAsXLSX(image_information: str|dict[str, Any]) -> bytes:
    import io as IO
    import openpyxl # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    wb = openpyxl.Workbook()
    data = json.loads(image_information) if isinstance(image_information, str) else image_information
    def add_sheet(sheet_name, data, col_widths=None, row_height=15):
        ws = wb.create_sheet(title=sheet_name)
        rows = [row.split('\t') for row in data.strip().split('\n')]

        for row_idx, row in enumerate(rows, start=1):
            ws.row_dimensions[row_idx].height = row_height
            for col_idx, cell in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=cell.replace('\n', '').replace('\r', ''))

        if isinstance(col_widths, list):
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        elif isinstance(col_widths, int):
            max_cols = max(len(row) for row in rows)
            for col_idx in range(1, max_cols + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths

    sections = {
        "Main Image Info": (export_main_image_info, [25, 100], 15),
        "Image Text Info": (export_image_text_info, [25, 100], 15),
        "Experiment Data": (export_experiments, 15, 15),
        "Recorded Data": (export_recorded_data, 15, 15),
        "Custom Metadata": (export_custom_metadata, [25, 100], 15),
        "Acquisition Details": (export_acquisition_details, 20, 15)
    }

    for sheet_name, (export_func, col_width, row_height) in sections.items():
        add_sheet(sheet_name, export_func(data), col_width, row_height)

    wb.remove(wb["Sheet"])
    result = IO.BytesIO()
    wb.save(result)
    result.seek(0)
    return result.read()

def format_file_size(size: int) -> str:
    kB = 1024
    MB = kB*1024
    GB = MB*1024
    TB = GB*1024
    if TB <= size:
        return f"{size/TB:.0f}TB"
    if GB <= size:
        return f"{size/GB:.0f}GB"
    if MB <= size:
        return f"{size/MB:.0f}MB"
    if kB <= size:
        return f"{size/kB:.0f}kB"
    return f"{size} B"


def format_general_info_sizes(file_bytes: int, frame_bytes: int, volume_bytes: int) -> dict[str, Any]:
    ret = {
        "file_bytes": file_bytes,
        "frame_bytes": frame_bytes,
        "volume_bytes": volume_bytes,
        "file_size": format_file_size(file_bytes),
        "frame_size": format_file_size(frame_bytes),
        "volume_size": format_file_size(volume_bytes)
    }
    ret["sizes"] = f"{ret['file_size']} on disk, {ret['frame_size']} frame" + (f", {ret['volume_size']} volume" if volume_bytes else "")
    return ret
